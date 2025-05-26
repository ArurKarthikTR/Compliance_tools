import os
import json
import csv
import io
import datetime
import random
import string
import pandas as pd
from flask import jsonify, send_file, send_from_directory, request
from openpyxl import Workbook

class TestGeneratorService:
    def __init__(self, download_folder):
        self.download_folder = download_folder
        os.makedirs(self.download_folder, exist_ok=True)
        
        # Import the Faker module
        try:
            from faker import Faker
            self.fake = Faker()
        except ImportError:
            # If Faker is not installed, we'll handle it later
            self.fake = None
    
    def generate_data(self, request_json):
        """
        Generate mock data based on the provided field configuration
        """
        try:
            # Check if Faker is available
            if self.fake is None:
                return jsonify({"error": "Faker module is not installed. Please install it with 'pip install faker'"}), 500
            
            fields = request_json.get('fields', [])
            row_count = request_json.get('rowCount', 10)
            
            # Validate input
            if not fields:
                return jsonify({"error": "No fields provided"}), 400
            
            if not isinstance(row_count, int) or row_count <= 0 or row_count > 1000:
                return jsonify({"error": "Invalid row count. Must be between 1 and 1000"}), 400
            
            # Generate data
            generated_data = self._generate_mock_data(fields, row_count)
            
            return jsonify(generated_data)
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    def download_csv(self, request_json):
        """
        Generate mock data and return it as a CSV file
        """
        try:
            # Check if Faker is available
            if self.fake is None:
                return jsonify({"error": "Faker module is not installed. Please install it with 'pip install faker'"}), 500
            
            fields = request_json.get('fields', [])
            row_count = request_json.get('rowCount', 10)
            
            # Validate input
            if not fields:
                return jsonify({"error": "No fields provided"}), 400
            
            if not isinstance(row_count, int) or row_count <= 0 or row_count > 1000:
                return jsonify({"error": "Invalid row count. Must be between 1 and 1000"}), 400
            
            # Generate data
            generated_data = self._generate_mock_data(fields, row_count)
            
            # Generate a unique filename
            filename = f"test-data-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
            filepath = os.path.join(self.download_folder, filename)
            
            # Save to file
            with open(filepath, 'w', newline='') as csvfile:
                if generated_data:
                    writer = csv.DictWriter(csvfile, fieldnames=generated_data[0].keys())
                    writer.writeheader()
                    writer.writerows(generated_data)
            
            # Create response
            return send_from_directory(
                self.download_folder, 
                filename, 
                as_attachment=True,
                mimetype='text/csv'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    def download_json(self, request_json):
        """
        Generate mock data and return it as a JSON file
        """
        try:
            # Check if Faker is available
            if self.fake is None:
                return jsonify({"error": "Faker module is not installed. Please install it with 'pip install faker'"}), 500
            
            fields = request_json.get('fields', [])
            row_count = request_json.get('rowCount', 10)
            
            # Validate input
            if not fields:
                return jsonify({"error": "No fields provided"}), 400
            
            if not isinstance(row_count, int) or row_count <= 0 or row_count > 1000:
                return jsonify({"error": "Invalid row count. Must be between 1 and 1000"}), 400
            
            # Generate data
            generated_data = self._generate_mock_data(fields, row_count)
            
            # Generate a unique filename
            filename = f"test-data-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.json"
            filepath = os.path.join(self.download_folder, filename)
            
            # Save to file
            with open(filepath, 'w') as jsonfile:
                json.dump(generated_data, jsonfile, indent=2)
            
            # Create response
            return send_from_directory(
                self.download_folder, 
                filename, 
                as_attachment=True,
                mimetype='application/json'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    def download_excel(self, request_json):
        """
        Generate mock data and return it as an Excel file
        """
        try:
            # Check if Faker is available
            if self.fake is None:
                return jsonify({"error": "Faker module is not installed. Please install it with 'pip install faker'"}), 500
            
            fields = request_json.get('fields', [])
            row_count = request_json.get('rowCount', 10)
            
            # Validate input
            if not fields:
                return jsonify({"error": "No fields provided"}), 400
            
            if not isinstance(row_count, int) or row_count <= 0 or row_count > 1000:
                return jsonify({"error": "Invalid row count. Must be between 1 and 1000"}), 400
            
            # Generate data
            generated_data = self._generate_mock_data(fields, row_count)
            
            # Generate a unique filename
            filename = f"test-data-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
            filepath = os.path.join(self.download_folder, filename)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Generated Data"
            
            if generated_data:
                # Add headers
                headers = list(generated_data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Add data rows
                for row_idx, row_data in enumerate(generated_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
            
            # Save the workbook
            wb.save(filepath)
            
            # Create response
            return send_from_directory(
                self.download_folder, 
                filename, 
                as_attachment=True,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
            
    def download_xml(self, request_json):
        """
        Generate mock data and return it as an XML file
        """
        try:
            # Check if Faker is available
            if self.fake is None:
                return jsonify({"error": "Faker module is not installed. Please install it with 'pip install faker'"}), 500
            
            fields = request_json.get('fields', [])
            row_count = request_json.get('rowCount', 10)
            
            # Validate input
            if not fields:
                return jsonify({"error": "No fields provided"}), 400
            
            if not isinstance(row_count, int) or row_count <= 0 or row_count > 1000:
                return jsonify({"error": "Invalid row count. Must be between 1 and 1000"}), 400
            
            # Generate data
            generated_data = self._generate_mock_data(fields, row_count)
            
            # Generate a unique filename
            filename = f"test-data-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xml"
            filepath = os.path.join(self.download_folder, filename)
            
            # Create XML content
            xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n<data>\n'
            
            for row in generated_data:
                xml_content += '  <row>\n'
                for key, value in row.items():
                    # Escape special characters in XML
                    if isinstance(value, str):
                        value = value.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&apos;')
                    xml_content += f'    <{key}>{value}</{key}>\n'
                xml_content += '  </row>\n'
            
            xml_content += '</data>'
            
            # Save to file
            with open(filepath, 'w', encoding='utf-8') as xmlfile:
                xmlfile.write(xml_content)
            
            # Create response
            return send_from_directory(
                self.download_folder, 
                filename, 
                as_attachment=True,
                mimetype='application/xml'
            )
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    def _generate_mock_data(self, fields, row_count):
        """
        Generate mock data based on field configurations
        
        Args:
            fields (list): List of field configurations
            row_count (int): Number of rows to generate
            
        Returns:
            list: List of dictionaries containing the generated data
        """
        generated_data = []
        unique_values = {}
        duplicate_values = {}
        
        # Initialize unique value tracking and generate duplicate values
        for field in fields:
            field_id = field['id']
            field_type = field['type']
            field_options = field.get('options', {})
            is_unique = field.get('unique', False)
            
            # Add field name to options for context-aware generation
            field_options['field_name'] = field['name']
            
            if is_unique:
                unique_values[field_id] = set()
            else:
                # For non-unique fields, generate a single value to use for all rows
                duplicate_values[field_id] = self._generate_field_value(field_type, field_options)
        
        # Generate data for each row
        for _ in range(row_count):
            row_data = {}
            
            for field in fields:
                field_id = field['id']
                field_type = field['type']
                field_options = field.get('options', {})
                is_unique = field.get('unique', False)
                
                # Add field name to options for context-aware generation
                field_options['field_name'] = field['name']
                
                if is_unique:
                    # Generate unique values
                    value = self._generate_field_value(field_type, field_options)
                    
                    # For date fields, ensure we generate truly unique dates
                    if field_type == 'Date' or field_type == 'date':
                        # Try up to 1000 times to generate a unique date
                        attempts = 0
                        max_attempts = 1000
                        
                        # Keep track of used dates to avoid duplicates
                        if field_id not in unique_values:
                            unique_values[field_id] = set()
                            
                        while str(value) in unique_values[field_id] and attempts < max_attempts:
                            # For dates, modify the options to ensure we get different dates
                            # Adjust the date range slightly for each attempt
                            if 'startDate' in field_options and 'endDate' in field_options:
                                try:
                                    start_date = datetime.datetime.strptime(field_options['startDate'], '%Y-%m-%d')
                                    end_date = datetime.datetime.strptime(field_options['endDate'], '%Y-%m-%d')
                                    
                                    # Adjust the date range slightly for each attempt
                                    adjusted_start = start_date - datetime.timedelta(days=attempts % 30)
                                    adjusted_end = end_date + datetime.timedelta(days=attempts % 30)
                                    
                                    field_options['startDate'] = adjusted_start.strftime('%Y-%m-%d')
                                    field_options['endDate'] = adjusted_end.strftime('%Y-%m-%d')
                                except (ValueError, TypeError):
                                    pass
                            
                            value = self._generate_field_value(field_type, field_options)
                            attempts += 1
                        
                        # If we still couldn't generate a unique date, use a date with offset
                        if attempts >= max_attempts:
                            try:
                                # Parse the date and add a specific number of days
                                date_obj = datetime.datetime.strptime(value, '%Y-%m-%d')
                                offset_days = len(unique_values[field_id]) + 1
                                new_date = date_obj + datetime.timedelta(days=offset_days)
                                value = new_date.strftime('%Y-%m-%d')
                            except (ValueError, TypeError):
                                # If parsing fails, add a suffix as a last resort
                                value = f"{value}_{len(unique_values[field_id])}"
                    else:
                        # For non-date fields, use the standard approach
                        attempts = 0
                        max_attempts = 100
                        while str(value) in unique_values[field_id] and attempts < max_attempts:
                            value = self._generate_field_value(field_type, field_options)
                            attempts += 1
                        
                        # If we couldn't generate a unique value after max attempts, add a suffix to make it unique
                        if attempts >= max_attempts:
                            value = f"{value}_{len(unique_values[field_id])}"
                    
                    unique_values[field_id].add(str(value))
                else:
                    # Use the same value for all rows (duplicate)
                    value = duplicate_values[field_id]
                
                row_data[field['name']] = value
            
            generated_data.append(row_data)
        
        return generated_data
    
    def _generate_field_value(self, field_type, options):
        """
        Generate a value for a specific field type
        
        Args:
            field_type (str): The type of field
            options (dict): Options for the field
            
        Returns:
            The generated value
        """
        # Only handle the allowed data types
        if field_type == 'String' or field_type == 'string':
            return self._generate_text(options)
        elif field_type == 'Number' or field_type == 'number':
            return self._generate_number(options)
        elif field_type == 'Float' or field_type == 'float':
            options['decimal'] = True
            return self._generate_number(options)
        elif field_type == 'Decimal' or field_type == 'decimal':
            options['decimal'] = True
            return self._generate_number(options)
        elif field_type == 'Boolean' or field_type == 'boolean':
            return self._generate_boolean(options)
        elif field_type == 'Date' or field_type == 'date':
            return self._generate_date(options)
        elif field_type == 'Constant' or field_type == 'constant':
            return self._generate_constant(options)
        else:
            return f"Unsupported field type: {field_type}"
            
    def _generate_constant(self, options):
        """Generate constant value"""
        # Simply return the constant value from options
        return options.get('value', '10')  # Default to '10' if no value is provided
    
    def _generate_text(self, options):
        """Generate text data"""
        min_length = options.get('minLength', 5)
        max_length = options.get('maxLength', 20)
        field_name = options.get('field_name', '').lower()
        
        if self.fake:
            # Generate data based on column name patterns
            # Check for common column name patterns and use appropriate Faker methods
            
            # Name related fields
            if any(name_pattern in field_name for name_pattern in ['name', 'fullname', 'full_name', 'person']):
                if 'first' in field_name or 'firstname' in field_name or 'fname' in field_name:
                    return self.fake.first_name()
                elif 'last' in field_name or 'lastname' in field_name or 'lname' in field_name:
                    return self.fake.last_name()
                else:
                    return self.fake.name()
            
            # Location related fields
            elif 'country' in field_name:
                return self.fake.country()
            elif 'city' in field_name:
                return self.fake.city()
            elif 'state' in field_name or 'province' in field_name:
                return self.fake.state()
            elif 'address' in field_name or 'street' in field_name:
                return self.fake.street_address()
            elif 'zip' in field_name or 'postal' in field_name:
                return self.fake.zipcode()
            
            # Contact information
            elif 'email' in field_name:
                return self.fake.email()
            elif 'phone' in field_name or 'mobile' in field_name or 'cell' in field_name:
                return self.fake.phone_number()
            
            # Company related fields
            elif 'company' in field_name or 'business' in field_name or 'organization' in field_name:
                return self.fake.company()
            elif 'job' in field_name or 'position' in field_name or 'title' in field_name:
                return self.fake.job()
            
            # Internet related fields
            elif 'username' in field_name or 'user_name' in field_name or 'userid' in field_name:
                return self.fake.user_name()
            elif 'url' in field_name or 'website' in field_name or 'web' in field_name:
                return self.fake.url()
            elif 'domain' in field_name:
                return self.fake.domain_name()
            elif 'ip' in field_name:
                return self.fake.ipv4()
            
            # Financial related fields
            elif 'credit' in field_name and 'card' in field_name:
                return self.fake.credit_card_number()
            elif 'currency' in field_name:
                return self.fake.currency_code()
            
            # Miscellaneous fields
            elif 'color' in field_name or 'colour' in field_name:
                return self.fake.color_name()
            elif 'product' in field_name or 'item' in field_name:
                return self.fake.catch_phrase()
            elif 'description' in field_name or 'desc' in field_name or 'comment' in field_name:
                return self.fake.text(max_nb_chars=max_length)[:max_length]
            
            # For very short strings, use more natural-looking text
            elif min_length < 5:
                # Try to use appropriate Faker methods based on length
                methods = [
                    self.fake.word,
                    self.fake.color_name,
                    self.fake.first_name,
                    self.fake.last_name,
                    self.fake.city_prefix,
                    self.fake.city_suffix
                ]
                
                # Try each method until we find one that generates text of appropriate length
                for method in methods:
                    for _ in range(5):  # Try a few times with each method
                        value = method()
                        if len(value) >= min_length and len(value) <= max_length:
                            return value
                
                # If we couldn't find a suitable value, use word() and adjust
                word = self.fake.word()
                if len(word) < min_length:
                    # Combine with another word instead of padding with repeated chars
                    second_word = self.fake.word()
                    combined = f"{word}{second_word}"
                    return combined[:max_length]
                elif len(word) > max_length:
                    return word[:max_length]
                return word
            else:
                # For normal length strings, use text() method
                generated_text = self.fake.text(max_nb_chars=max_length)
                # Trim to max_length if needed
                if len(generated_text) > max_length:
                    return generated_text[:max_length]
                # Pad to min_length if needed
                elif len(generated_text) < min_length:
                    # Pad with another word instead of spaces
                    additional_text = self.fake.word()
                    return (generated_text + " " + additional_text)[:max_length]
                return generated_text
        else:
            # Fallback if Faker is not available
            length = random.randint(min_length, max_length)
            return ''.join(random.choice(string.ascii_letters + ' ') for _ in range(length))
    
    def _generate_number(self, options):
        """Generate number data"""
        min_value = options.get('min', 0)
        max_value = options.get('max', 100)
        decimal = options.get('decimal', False)
        length = options.get('length')
        field_name = options.get('field_name', '').lower()
        
        # Generate different ranges based on field name to create more diverse data
        if 'code' in field_name:
            # For code fields, generate codes with specific patterns
            if 'entity' in field_name:
                # Entity codes: E-1000 to E-9999
                return random.randint(1000, 9999)
            elif 'post' in field_name or 'zip' in field_name:
                # Postal codes: 5-digit numbers
                return random.randint(10000, 99999)
            else:
                # Other codes: 100-999
                return random.randint(100, 999)
        elif 'id' in field_name:
            # IDs: 1000-9999
            return random.randint(1000, 9999)
        elif 'vat' in field_name:
            # VAT numbers: 10000-99999
            return random.randint(10000, 99999)
        elif 'year' in field_name:
            # Years: 2020-2025
            return random.randint(2020, 2025)
        elif 'value' in field_name or 'amount' in field_name:
            # Values/amounts: 100-10000
            if decimal:
                return round(random.uniform(100, 10000), 2)
            else:
                return random.randint(100, 10000)
        elif decimal:
            return round(random.uniform(min_value, max_value), 2)
        elif length:
            # Generate a number with specific length
            length = int(length)
            if length <= 0:
                return random.randint(min_value, max_value)
            
            # Generate a random number with exactly 'length' digits
            if length == 1:
                return random.randint(0, 9)
            else:
                min_length_value = 10 ** (length - 1)
                max_length_value = (10 ** length) - 1
                
                # Respect min and max constraints
                actual_min = max(min_value, min_length_value)
                actual_max = min(max_value, max_length_value)
                
                # If constraints conflict, prioritize length
                if actual_min > actual_max:
                    actual_min = min_length_value
                    actual_max = max_length_value
                
                return random.randint(actual_min, actual_max)
        else:
            return random.randint(min_value, max_value)
    
    def _generate_date(self, options):
        """Generate date data"""
        # Default dates - use more recent years (last 5 years to next 2 years)
        current_year = datetime.datetime.now().year
        default_start = datetime.datetime(current_year - 5, 1, 1)
        default_end = datetime.datetime(current_year + 2, 12, 31)
        
        # Get date strings from options
        start_date_str = options.get('startDate', f'{current_year-5}-01-01')
        end_date_str = options.get('endDate', f'{current_year+2}-12-31')
        field_name = options.get('field_name', '').lower()
        
        # Try to parse the dates
        try:
            start = datetime.datetime.strptime(start_date_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            start = default_start
            
        try:
            end = datetime.datetime.strptime(end_date_str, '%Y-%m-%d')
        except (ValueError, TypeError):
            end = default_end
            
        # Ensure end date is not before start date
        if end < start:
            end = start + datetime.timedelta(days=365)  # Default to 1 year range
        
        # Adjust date ranges based on field name
        if 'birth' in field_name or 'dob' in field_name:
            # Birth dates should be older (18-80 years ago)
            years_ago = random.randint(18, 80)
            start = datetime.datetime.now() - datetime.timedelta(days=365*years_ago)
            end = datetime.datetime.now() - datetime.timedelta(days=365*18)
        elif 'future' in field_name or 'target' in field_name or 'deadline' in field_name:
            # Future dates (next 1-5 years)
            start = datetime.datetime.now()
            end = datetime.datetime.now() + datetime.timedelta(days=365*5)
        elif 'expiry' in field_name or 'expiration' in field_name:
            # Expiry dates (next 1-3 years)
            start = datetime.datetime.now()
            end = datetime.datetime.now() + datetime.timedelta(days=365*3)
        
        if self.fake:
            return self.fake.date_between(start_date=start, end_date=end).strftime('%Y-%m-%d')
        else:
            # Simple random date between start and end
            delta = end - start
            random_days = random.randint(0, max(0, delta.days))
            return (start + datetime.timedelta(days=random_days)).strftime('%Y-%m-%d')
    
    def _generate_boolean(self, options):
        """Generate boolean data"""
        return random.choice([True, False])
    
    # Only keeping the methods needed for the supported data types
    
    def import_file(self, request):
        """
        Import a CSV or Excel file and extract column information
        
        Args:
            request: Flask request object containing the file
            
        Returns:
            JSON response with field configurations
        """
        try:
            # Check if file is present in the request
            if 'file' not in request.files:
                return jsonify({"error": "No file provided"}), 400
                
            file = request.files['file']
            
            # Check if filename is empty
            if file.filename == '':
                return jsonify({"error": "No file selected"}), 400
                
            # Check file extension
            if not (file.filename.endswith('.csv') or file.filename.endswith('.xlsx')):
                return jsonify({"error": "Only CSV and Excel files are supported"}), 400
            
            # Read the file using pandas
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:  # Excel file
                df = pd.read_excel(file)
            
            # Check if the file has columns (headers)
            if df.empty and len(df.columns) == 0:
                return jsonify({"error": "The file is completely empty (no headers)"}), 400
                
            # Even if there's no data rows but there are column headers, we can still process it
                
            # Extract column information and detect data types
            fields = []
            for column in df.columns:
                field_type = self._detect_column_type(df[column])
                field_options = self._generate_options_for_type(field_type, df[column])
                
                field = {
                    "name": column,
                    "type": field_type,
                    "options": field_options,
                    "unique": True  # Default to unique values
                }
                fields.append(field)
            
            return jsonify({"fields": fields})
        except Exception as e:
            return jsonify({"error": str(e)}), 500
    
    def _detect_type_from_column_name(self, column_name_lower):
        """
        Detect data type based on column name patterns
        
        Args:
            column_name_lower: Lowercase column name
            
        Returns:
            str: The detected data type or "Unknown" if no pattern matches
        """
        # Number type patterns
        number_keywords = [
            'code', 'id', 'number', 'num', 'entity', 'vat', 'zip', 'sales_id', 
            'year', 'count', 'quantity', 'qty', 'amount', 'total', 'value', 'price', 
            'cost', 'invoice', 'order', 'item', 'product', 'sku', 'stock', 'inventory'
        ]
        
        # Date type patterns
        date_keywords = [
            'date', 'time', 'day', 'month', 'year', 'created', 'modified', 'updated',
            'start', 'end', 'begin', 'finish', 'due', 'deadline', 'scheduled', 'reporting'
        ]
        
        # Boolean type patterns
        boolean_keywords = [
            'is', 'has', 'can', 'should', 'will', 'active', 'enabled', 'flag', 'status',
            'approved', 'confirmed', 'verified', 'completed', 'finished', 'done'
        ]
        
        # Float/Decimal type patterns
        float_keywords = [
            'price', 'amount', 'total', 'cost', 'fee', 'rate', 'percentage', 'percent',
            'decimal', 'float', 'double', 'average', 'mean', 'median', 'ratio'
        ]
        
        # Special case for country names
        if 'country' in column_name_lower:
            return "String"
            
        # Special case for postal/zip codes
        if 'post' in column_name_lower or 'zip' in column_name_lower:
            return "Number"
            
        # Special case for VAT numbers
        if 'vat' in column_name_lower:
            return "Number"
            
        # Special case for entity codes
        if 'entity' in column_name_lower and ('code' in column_name_lower or 'id' in column_name_lower):
            return "Number"
            
        # Check for date patterns
        if any(keyword in column_name_lower for keyword in date_keywords):
            return "Date"
            
        # Check for number patterns
        if any(keyword in column_name_lower for keyword in number_keywords):
            return "Number"
            
        # Check for boolean patterns
        if any(keyword in column_name_lower for keyword in boolean_keywords) and len(column_name_lower) < 15:
            return "Boolean"
            
        # Check for float patterns
        if any(keyword in column_name_lower for keyword in float_keywords):
            return "Float"
            
        # If no pattern matches, return Unknown
        return "Unknown"
    
    def _detect_column_type(self, column):
        """
        Detect the data type of a column
        
        Args:
            column: Pandas Series representing a column
            
        Returns:
            str: The detected data type
        """
        # Get column name for pattern matching
        column_name = column.name if hasattr(column, 'name') else ""
        column_name_lower = str(column_name).lower()
        
        # Drop NaN values for analysis
        clean_column = column.dropna()
        
        # For empty columns, rely heavily on column name patterns
        if clean_column.empty:
            # Check column name patterns for type detection
            type_from_name = self._detect_type_from_column_name(column_name_lower)
            return "String" if type_from_name == "Unknown" else type_from_name
        
        # Check for boolean type
        if column.dtype == bool or (clean_column.isin([0, 1, True, False]).all() and len(clean_column.unique()) <= 2):
            return "Boolean"
        
        # Column name hints for specific types - check these first before analyzing data
        type_from_name = self._detect_type_from_column_name(column_name_lower)
        if type_from_name != "Unknown":
            return type_from_name
        
        # Check for date type if it's already a datetime type
        if column.dtype == 'datetime64[ns]':
            try:
                dates = pd.to_datetime(clean_column, errors='coerce')
                if dates.notna().sum() > 0.7 * len(clean_column):  # If >70% are valid dates
                    return "Date"
            except:
                pass
        
        # Check for numeric types
        if pd.api.types.is_integer_dtype(column):
            return "Number"
        
        if pd.api.types.is_float_dtype(column):
            # Check if all values have decimal places
            try:
                if all(x == int(x) for x in clean_column):
                    return "Number"
                return "Float"
            except:
                pass
        
        # Try to convert to numeric
        try:
            numeric_column = pd.to_numeric(clean_column, errors='coerce')
            if numeric_column.notna().sum() > 0.7 * len(clean_column):  # If >70% are valid numbers
                # Check if all values are integers
                if all(float(x).is_integer() for x in numeric_column.dropna()):
                    return "Number"
                return "Float"
        except:
            pass
        
        # Default to string
        return "String"
    
    def _generate_options_for_type(self, field_type, column):
        """
        Generate options for a field based on its type and data
        
        Args:
            field_type: The detected field type
            column: Pandas Series representing a column
            
        Returns:
            dict: Options for the field
        """
        options = {}
        
        if field_type == "String":
            # Get min and max length of strings
            lengths = column.astype(str).str.len()
            options["minLength"] = int(max(1, lengths.min()))
            options["maxLength"] = int(min(100, lengths.max() * 1.2))  # Add 20% buffer
            
        elif field_type == "Number":
            # Get min and max values, handling NaN values
            clean_column = column.dropna()
            if clean_column.empty:
                options["min"] = 0
                options["max"] = 100
                options["length"] = 3
            else:
                # Use safe conversion to int with fallback
                try:
                    min_val = int(clean_column.min())
                except (ValueError, TypeError):
                    min_val = 0
                
                try:
                    max_val = int(clean_column.max())
                except (ValueError, TypeError):
                    max_val = 100
                
                options["min"] = min_val
                options["max"] = max_val
                
                # Get max length safely
                try:
                    max_length = len(str(int(clean_column.max())))
                except (ValueError, TypeError):
                    max_length = 3
                
                options["length"] = max_length
            
        elif field_type == "Float" or field_type == "Decimal":
            # Get min and max values, handling NaN values
            clean_column = column.dropna()
            if clean_column.empty:
                options["min"] = 0.0
                options["max"] = 100.0
                options["precision"] = 2
            else:
                # Use safe conversion to float with fallback
                try:
                    min_val = float(clean_column.min())
                except (ValueError, TypeError):
                    min_val = 0.0
                
                try:
                    max_val = float(clean_column.max())
                except (ValueError, TypeError):
                    max_val = 100.0
                
                options["min"] = min_val
                options["max"] = max_val
                
                # Get precision (number of decimal places)
                max_decimals = 0
                for val in clean_column:
                    try:
                        decimals = len(str(val).split('.')[-1]) if '.' in str(val) else 0
                        max_decimals = max(max_decimals, decimals)
                    except:
                        pass
                
                options["precision"] = max_decimals
            
        elif field_type == "Date":
            # Convert to datetime if not already
            if column.dtype != 'datetime64[ns]':
                column = pd.to_datetime(column, errors='coerce')
            
            # Get min and max dates, handling NaT values
            clean_column = column.dropna()
            if clean_column.empty:
                # Default date range if all values are NaT
                options["startDate"] = '2020-01-01'
                options["endDate"] = '2025-12-31'
            else:
                min_date = clean_column.min()
                max_date = clean_column.max()
                
                options["startDate"] = min_date.strftime('%Y-%m-%d')
                options["endDate"] = max_date.strftime('%Y-%m-%d')
        
        return options
